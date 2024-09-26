#!/user/bin/env python
# coding=utf-8
import time
import shutil
import pandas as pd
import os
from openpyxl import *
import re
import jieba
import jieba.analyse
import numpy as np


# 统一label
def unify_labels(path1, filenames):
    for i in range(len(filenames)):
        excel_path = path1 + filenames[i]
        print(excel_path)
        wb = load_workbook(excel_path)
        ws = wb.active
        ws.delete_rows(1)
        ws.insert_rows(1)
        ws['A1'] = ''
        ws['B1'] = 'rank'
        ws['C1'] = 'title'
        ws['D1'] = 'url'
        ws['E1'] = 'heat'
        wb.save(excel_path)


# 提取并保存words
docs_del_inix = []
def get_save_words(data, f_name, words_p):
    content = data
    temp = re.sub('[a-zA-Z]', ' ', content)  # 去掉英文

    remove_chars = '[0-9’!"#$%&\'()*+,-./:;<=>?@，。?★、…【】《》？“”‘’！[\\]^_`{|}~]+'
    # remove_chars = '[’!"#$%&\'()*+,-./:;<=>?@，。?★、…【】《》？“”‘’！[\\]^_`{|}~]+'

    temp = re.sub(remove_chars, '', temp)
    tags = jieba.analyse.extract_tags(temp, topK=12)
    contents = ",".join(tags)
    path = words_p + f_name
    # 剔除掉词个数太少的数据
    if len(tags) < 6 or len(temp) == 0:
        docs_del_inix.append(f_name)
    else:
        f = open(path, 'w', encoding='utf-8')
        f.write(contents)
        f.close()
    return docs_del_inix

# 保存数据到txt
def save_data(title, heat, meatdata, abs, path, heat_lst):
    heat_lst.append(heat)
    f_heat = open('/heat.txt', 'a+', encoding='utf-8')
    f_heat.write(str(heat))
    f_heat.write('\n')
    f = open(path, 'w', encoding='utf-8')
    # f.write('heat:')
    # f.write(str(heat))
    # f.write("++++")
    # f.write('title:')
    f.write(title)
    f.write(' ')

    # f.write(' ')
    for i in range(len(meatdata)):
        # f.write('metadata:')
        # f.write(' ')
        # f.write(str(meatdata[i]))
        # f.write('abs:')
        f.write(str(abs[i]))
        f.write(';')
    f.close()
    f_heat.close()
    return heat_lst


# 保存带有热度值的数据

def get_save_datawithheat(f_name, data_txt, data_withheat_path):
    f = open(data_withheat_path + f_name, 'w', encoding='utf-8')
    f.write(data_txt)
    f.close()


# load excel数据
def load_data(path1, path2, txt_p, words_p, path3, path4, path5):
    heat_lst = []
    filenames = os.listdir(path1)
    for i in range(len(filenames)):
        # 从filenames1中得到title和heat
        excel1_path = path1 + filenames[i]
        # print('excel1_name:',excel1_path)
        f1 = open(excel1_path, 'rb')
        data1 = pd.read_excel(f1)
        title = data1['title']
        heat = data1['heat']

        excel2_path = path2 + filenames[i]
        # print('excel2_name:',excel2_path)
        print('正在读取文件{:d}'.format(i))
        f2 = open(excel2_path, 'rb')
        data2 = pd.read_excel(f2)
        metadata = data2['meta']
        abs = data2['abs']
        title_temp = data2['title']
        metadata_lst = []
        abs_lst = []
        count = 0

        for j in range(len(title_temp)-1):
            if title_temp[j] == '\n':
                title1 = title[count]
                heat1 = heat[count]
                txt_name = filenames[i].replace('.xlsx', '.txt')
                temp = list(txt_name)
                temp.insert(-4, '-')
                temp.insert(-4, count)
                txt_name = ''.join(str(i) for i in temp)
                txt_path = txt_p + txt_name
                txt_data = str('title:') + title1 + str(' heat:') + str(heat1) + str(' ') + "".join(abs_lst)
                txt_title = title1
                # txt_data_heat = str(heat1) + str('++++') + "".join(abs_lst)
                txt_data_heat = str(heat1)
                docs_del_inix = get_save_words(txt_data, txt_name, words_p)
                # 保存带有热度的txt_data
                get_save_datawithheat(txt_name, txt_data_heat, path4)
                # 保存带有热度和名字的txt_data
                get_save_datawithheat(txt_name, txt_data, path3)
                # 保存带有名字的txt_title
                get_save_datawithheat(txt_name, txt_title, path5)
                heat_lst = save_data(title1, heat1, metadata_lst, abs_lst, txt_path, heat_lst)
                metadata_lst = []
                abs_lst = []
                print('写入文件{:d}成功'.format(count))
                count = count + 1
                continue
            else:
                if isinstance(abs[j], str):
                    metadata_lst.append(metadata[j])
                else:
                    metadata_lst.append(str(' '))
                if isinstance(abs[j], str):
                    abs_lst.append(abs[j])
                else:
                    abs_lst.append(str(' '))
    # return heat_lst
    return docs_del_inix

# 整合所有txt到一个txt文件
def MergeTxt(filepath, outfile):
    k = open(outfile, 'a+', encoding='utf-8')
    for parent, dirnames, filenames in os.walk(filepath):
        for filepath in filenames:
            txtPath = os.path.join(parent, filepath)  # txtpath就是所有文件夹的路径
            f = open(txtPath, 'a+', encoding='utf-8')
            f.seek(0)
            str = f.read()
            k.write(str + "\n")
    k.close()
    print("finished")


def main():
    """
    生成 word2vecter 预训练文本数据

    """

    # 文件路径
    # path1 = 'E:\\JLL-work\\crawler_timeseries\\data_baidu_time-2022\\'
    # path2 = 'E:\\JLL-work\\crawler_timeseries\\data-baidu-title-three-2022\\'
    # txt_path1 = 'E:\\JLL-work\\code\\data\\baidu_txt\\'
    # txt_path3 = "E:\\JLL-work\\code\\data\\baidu_txt_heat_title\\"
    # txt_path2 = 'E:\\JLL-work\\code\\data\\baidu_txt_heat\\'
    # txt_words_path = 'E:\\JLL-work\\code\\data\\baidu_txt_words\\'
    # heat_lst = load_data(path1, path2, txt_path1, txt_words_path, txt_path3, txt_path2)
    # heat = np.array(heat_lst)
    # np.save("heat_lst", heat)

    """
    生成模型用于训练的文本数据
    
    """

    path1 = 'E:\\LLL-work\\crawler\\data_noTime_all\\data_baidu_time_new-2022\\'
    path2 = 'E:\\LLL-work\\crawler\\data_noTime_all\\data-baidu-title-three-2022\\'
    txt_path1 = 'E:\\JLL-work\\code\\data\\all_txt\\'
    txt_path2 = 'E:\\JLL-work\\code\\data\\all_heat\\'
    txt_path3 = "E:\\JLL-work\\code\\data\\all_txt_heat_title\\"
    txt_path4 = 'E:\\JLL-work\\code\\data\\all_title\\'
    txt_words_path = 'E:\\JLL-work\\code\\data\\all_txt_words\\'
    docs_del_inix = load_data(path1, path2, txt_path1, txt_words_path, txt_path3, txt_path2, txt_path4)

    return docs_del_inix

# def main_noTime():
#     # 文件路径
#     path1 = 'E:\\LLL-work\\crawler\\data_noTime_train\\data_baidu_time\\'
#     filenames = os.listdir(path1)
#     # unify_labels(path1,filenames1)
#     path2 = 'E:\\LLL-work\\crawler\\data_noTime_train\\data-baidu-title\\'
#     # filenames2 = os.listdir(path2)
#     txt_path1 = 'E:\\JLL-work\\code\\noTime_data\\baidu_txt\\'
#     txt_path2 = 'E:\\JLL-work\\code\\noTime_data\\baidu_txt_heat\\'
#     txt_path3 = 'E:\\JLL-work\\code\\noTime_data\\baidu_txt_heat_title\\'
#     txt_words_path = 'E:\\JLL-work\\code\\noTime_data\\baidu_txt_words\\'
#     heat_lst = load_data(path1, filenames, path2, txt_path1, txt_words_path, txt_path3, txt_path2)
#     heat = np.array(heat_lst)
#     np.save("heat_lst", heat)


def get_one():
    """
    整合txt文件用于预训练词向量模型

    """

    # path_txt1 = "E:\\JLL-work\\code\\data\\baidu_txt"
    # outfile_alltxt1 = "E:\\JLL-work\\code\\data\\baidu_txt_all_lines.txt"
    # MergeTxt(path_txt1, outfile_alltxt1)

    # path_txt2 = "E:\\JLL-work\\code\\data\\baidu_txt_heat"
    # outfile_alltxt2 = "E:\\JLL-work\\code\\data\\baidu_txt_heat_all.txt"
    # MergeTxt(path_txt2, outfile_alltxt2)

    # path_txt3 = "E:\\JLL-work\\code\\data\\baidu_txt_heat_title"
    # outfile_alltxt3 = "E:\\JLL-work\\code\\data\\baidu_txt_heat_title_all.txt"


    """
    整合txt文件用于训练的文本数据

    """
    # path_txt1 = "E:\\JLL-work\\code\\data\\all_txt"
    # outfile_alltxt1 = "E:\\JLL-work\\code\\data\\data_txt_lines.txt"
    # MergeTxt(path_txt1, outfile_alltxt1)
    #
    # path_txt2 = "E:\\JLL-work\\code\\data\\all_heat"
    # outfile_alltxt2 = "E:\\JLL-work\\code\\data\\data_heat.txt"
    # MergeTxt(path_txt2, outfile_alltxt2)

    path_txt3 = "E:\\JLL-work\\code\\data\\all_title"
    outfile_alltxt3 = "E:\\JLL-work\\code\\data\\data_title.txt"
    MergeTxt(path_txt3, outfile_alltxt3)

def select_news(alltxt_path, save_path):
    filenames = os.listdir(alltxt_path)
    for i in range(len(filenames)):
        f = open(alltxt_path + filenames[i], 'r+', encoding='utf-8')
        line = f.readline()

        temp_title = re.findall(r'title:(.*?) heat', line)[0]  # 找到新闻题目
        str = filenames[i][11:27]  # 挑选出新闻发生时间

        path = save_path + temp_title + '.txt'

        # 文档名字不包含特殊字符
        intab = "?*|><"
        outtab = "     "
        trantab = str.maketrans(intab, outtab)
        path_t = path.translate(trantab)

        select_news = open(path_t, 'a+', encoding='utf-8')
        content = str + ' ' + line + "\n"
        select_news.write(content)

    select_news.close()


# 挑选具有一定出现频率的新闻
def select_fre(select_path, frequence, select_fre_path):
    filenames_fre = os.listdir(select_path)

    # 移动目标文件夹的根目录
    movabs_path = select_fre_path
    # 移动文件夹的根目录
    rawabs_path = select_path

    for i in range(len(filenames_fre)):
        f = open((select_path + filenames_fre[i]), 'r', encoding='utf-8')
        txt = f.read()
        count_fre = txt.count("title")  # 计算一个文档中相同新闻的个数

        if count_fre >= frequence:
            # 移动操作
            shutil.copy(rawabs_path + filenames_fre[i], movabs_path + filenames_fre[i])

    f.close()

# 删除不合格的文档
def del_docs(docs_del_inix):
    # path1 = 'E:\\JLL-work\\code\\data\\all_txt\\'
    # path2 = 'E:\\JLL-work\\code\\data\\all_heat\\'
    # path3 = 'E:\\JLL-work\\code\\data\\all_txt_heat_title\\'
    path4 = 'E:\\JLL-work\\code\\data\\all_title\\'

    for i in range(len(docs_del_inix)):
        # os.remove(os.path.join(path1, docs_del_inix[i]))
        # os.remove(os.path.join(path2, docs_del_inix[i]))
        # os.remove(os.path.join(path3, docs_del_inix[i]))
        os.remove(os.path.join(path4, docs_del_inix[i]))


if __name__ == '__main__':
    docs_del_indix = main()  # 步骤1
    docs_del_indix = np.array(docs_del_indix)
    # np.save('docs_del_indix.npy', docs_del_indix)

    docs_del_indix = np.load('docs_del_indix.npy')

    # # 删除不合格(一件事的相关词少于12个)的文档
    del_docs(docs_del_indix)  # 步骤2

    get_one()  # 步骤3

    time.time()

    # # 挑选相同的新闻放进新的文档
    # alltxt_path = 'E:\\JLL-work\\code\\data\\baidu_txt_heat_title\\'
    # select_path = 'E:\\JLL-work\\code\\data\\baidu_txt_heat_title_all_select\\'
    #
    # # select_news(alltxt_path, select_path)  # 步骤4
    #
    # # 设置新闻出现频率 筛选数据
    # fre = 24
    # select_fre_path = 'E:\\JLL-work\\code\\data\\baidu_txt_heat_title_all_select_frequence\\'
    # select_fre(select_path, fre, select_fre_path)  # 步骤5
